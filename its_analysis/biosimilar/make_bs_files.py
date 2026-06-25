"""
NDB注射月別データ から BS（バイオシミラー）行をハイライト・抽出するスクリプト
対象ファイル:
  - rd10_injection_monthly.xlsx  (No.10 FY2023)
  - No.11 FY2024 注射月別データ
"""

import openpyxl
from openpyxl.styles import PatternFill
from copy import copy
import os

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

RD11_SRC = "its_analysis/rd11_injection_monthly.xlsx"

FILES = [
    {
        "label": "rd10",
        "src": "/mnt/c/Users/e231b/WorkingDir/ndb_analysis/rd10_injection.xlsx",
        "out_hl": "its_analysis/biosimilar/rd10_injection_monthly_highlighted.xlsx",
        "out_bs": "its_analysis/biosimilar/rd10_injection_monthly_BS_only_v2.xlsx",
    },
    {
        "label": "rd11",
        "src": RD11_SRC,
        "out_hl": "its_analysis/biosimilar/rd11_injection_monthly_highlighted.xlsx",
        "out_bs": "its_analysis/biosimilar/rd11_injection_monthly_BS_only_v2.xlsx",
    },
]

# 341 サブラッド血液ろ過用補充液BSGは製品名に「ＢＳ」を含むがバイオシミラーではないため除外
EXCLUDE_DRUG_NAME_PATTERNS = ["サブラッド"]

HEADER_ROWS = 4  # 行1:タイトル, 行2:空白, 行3:列ヘッダー, 行4:空白


def is_bs(row_values):
    """医薬品名（列D=index3）に ＢＳ を含み、かつ除外リストにないか"""
    name = row_values[3]
    if name is None:
        return False
    name_str = str(name)
    if "ＢＳ" not in name_str:
        return False
    # サブラッドBSGはバイオシミラーではないため除外
    if any(pat in name_str for pat in EXCLUDE_DRUG_NAME_PATTERNS):
        return False
    return True


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font      = copy(src_cell.font)
        dst_cell.border    = copy(src_cell.border)
        dst_cell.fill      = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment  = copy(src_cell.alignment)


def process_file(cfg):
    label   = cfg["label"]
    src     = cfg["src"]
    out_hl  = cfg["out_hl"]
    out_bs  = cfg["out_bs"]

    print(f"\n{'='*60}")
    print(f"[{label}] {os.path.basename(src)}")

    wb_src = openpyxl.load_workbook(src)

    # ── ① ハイライト版 ────────────────────────────
    wb_hl = openpyxl.load_workbook(src)
    for sheet_name in wb_hl.sheetnames:
        ws = wb_hl[sheet_name]
        for row in ws.iter_rows(min_row=HEADER_ROWS + 1):
            vals = tuple(c.value for c in row)
            if is_bs(vals):
                for cell in row:
                    cell.fill = YELLOW
    wb_hl.save(out_hl)
    print(f"  ✓ ハイライト版保存: {out_hl}")

    # ── ② BS抽出版（列A・B 前方補完） ──────────────
    wb_bs = openpyxl.Workbook()
    wb_bs.remove(wb_bs.active)  # デフォルトシートを削除

    for sheet_name in wb_src.sheetnames:
        ws_src = wb_src[sheet_name]
        all_rows = list(ws_src.iter_rows(values_only=True))

        # 前方補完しながらデータ行を走査
        current_cat      = None
        current_cat_name = None
        bs_data_rows = []

        for r in all_rows[HEADER_ROWS:]:
            # 列A・B 前方補完
            if r[0] is not None:
                current_cat      = r[0]
                current_cat_name = r[1]
            filled_row = (current_cat, current_cat_name) + r[2:]

            if is_bs(filled_row):
                bs_data_rows.append(filled_row)

        # 新シートに書き込み
        ws_out = wb_bs.create_sheet(title=sheet_name)

        # ヘッダー行（行1〜4）をそのままコピー
        for r_idx, r_vals in enumerate(all_rows[:HEADER_ROWS], start=1):
            for c_idx, val in enumerate(r_vals, start=1):
                ws_out.cell(row=r_idx, column=c_idx, value=val)

        # BS行を続けて書き込み
        for r_idx, r_vals in enumerate(bs_data_rows, start=HEADER_ROWS + 1):
            for c_idx, val in enumerate(r_vals, start=1):
                ws_out.cell(row=r_idx, column=c_idx, value=val)

        bs_count = len(bs_data_rows)
        cats = sorted(set(r[0] for r in bs_data_rows if r[0] is not None))
        print(f"  [{sheet_name}] BS行数: {bs_count}, 薬効分類: {cats}")

    wb_bs.save(out_bs)
    print(f"  ✓ BS抽出版保存: {out_bs}")


if __name__ == "__main__":
    for cfg in FILES:
        process_file(cfg)
    print("\n全処理完了")
