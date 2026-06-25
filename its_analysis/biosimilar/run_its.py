"""
バイオシミラー注射薬 ITS（中断時系列）解析
===========================================
データ  : NDBオープンデータ 注射薬 診療月別 BS抽出版
          - rd10 (FY2023: 2023年4月〜2024年3月)
          - rd11 (FY2024: 2024年4月〜2025年3月)
介入    : 2024年4月1日（令和6年度診療報酬改定）
          → 入院患者へのBS使用促進評価新設
アウトカム: 月次BS処方数量（人口100,000人あたり）
設定    : 外来（院内+院外合算） / 入院  ×  9薬効分類 = 18モデル

人口データ出典: 総務省 人口推計（e-Stat 参考表 全国人口の推移）
  - 2024年・2025年: 月次確定値を使用
  - 2023年: 2022年10月(124,946,789)〜2023年10月(124,351,877)を線形補間
             2023年11月〜12月: 2023年10月〜2024年1月を線形補間
"""

import warnings
warnings.filterwarnings("ignore")

import pandas as pd
import numpy as np
import openpyxl
import statsmodels.api as sm
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from pathlib import Path

# ────────────────────────────────────────────────────────────────
# 定数
# ────────────────────────────────────────────────────────────────

BS_ONLY_FILES = {
    "rd10": "its_analysis/biosimilar/rd10_injection_monthly_BS_only_v2.xlsx",
    "rd11": "its_analysis/biosimilar/rd11_injection_monthly_BS_only_v2.xlsx",
}

# FY2023 の月列名 → 年月 (4月=2023-04 ... 3月=2024-03)
FY2023_MONTH_MAP = {
    "4月": "2023-04", "5月": "2023-05", "6月": "2023-06",
    "7月": "2023-07", "8月": "2023-08", "9月": "2023-09",
    "10月": "2023-10", "11月": "2023-11", "12月": "2023-12",
    "1月": "2024-01", "2月": "2024-02", "3月": "2024-03",
}
FY2024_MONTH_MAP = {
    "4月": "2024-04", "5月": "2024-05", "6月": "2024-06",
    "7月": "2024-07", "8月": "2024-08", "9月": "2024-09",
    "10月": "2024-10", "11月": "2024-11", "12月": "2024-12",
    "1月": "2025-01", "2月": "2025-02", "3月": "2025-03",
}

# 薬効分類名称
CAT_NAMES = {
    131: "眼科用剤",
    239: "消化器官用薬",
    241: "脳下垂体ホルモン剤",
    243: "甲状腺・副甲状腺ホルモン剤",
    249: "その他ホルモン剤",
    339: "その他血液・体液用薬",
    395: "酵素製剤",
    399: "代謝性医薬品(他分類外)",
    429: "その他腫瘍用薬",
}

INTERVENTION_MONTH = "2024-04"  # 介入点
HEADER_ROWS = 4

OUTPUT_DIR = Path("its_analysis/biosimilar/results")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ────────────────────────────────────────────────────────────────
# 人口データ（総務省 人口推計）
# ────────────────────────────────────────────────────────────────

def build_population_series() -> pd.Series:
    """
    2023年4月〜2025年3月の月別総人口（人）を返す。

    出典: 総務省 人口推計 e-Stat 参考表（全国人口の推移）
      - 2024年1月〜2025年3月 : 月次確定値
      - 2023年4月〜12月     : 2022年10月 / 2023年10月 の年次値から線形補間

    注: 各月の人口は「当該月の1日現在人口」
    """
    # 確定値（月次）
    confirmed = {
        "2024-01": 124_143_128, "2024-02": 124_105_194, "2024-03": 124_002_734,
        "2024-04": 124_001_809, "2024-05": 123_941_300, "2024-06": 123_979_444,
        "2024-07": 123_975_371, "2024-08": 123_887_489, "2024-09": 123_778_791,
        "2024-10": 123_801_750, "2024-11": 123_784_095, "2024-12": 123_744_278,
        "2025-01": 123_551_595, "2025-02": 123_441_499, "2025-03": 123_420_055,
    }

    # 2023年4月〜12月 の線形補間
    # 基点: 2022年10月1日 = 124,946,789  /  2023年10月1日 = 124,351,877
    pop_2022oct = 124_946_789
    pop_2023oct = 124_351_877
    delta_per_month_a = (pop_2023oct - pop_2022oct) / 12  # Oct22→Oct23

    # Oct22を0として各月オフセット
    months_2023 = {
        "2023-04": pop_2022oct + 6  * delta_per_month_a,
        "2023-05": pop_2022oct + 7  * delta_per_month_a,
        "2023-06": pop_2022oct + 8  * delta_per_month_a,
        "2023-07": pop_2022oct + 9  * delta_per_month_a,
        "2023-08": pop_2022oct + 10 * delta_per_month_a,
        "2023-09": pop_2022oct + 11 * delta_per_month_a,
        "2023-10": pop_2023oct,
    }
    # 2023年11月〜12月: Oct23 → Jan24 の線形補間
    delta_per_month_b = (confirmed["2024-01"] - pop_2023oct) / 3
    months_2023["2023-11"] = pop_2023oct + delta_per_month_b
    months_2023["2023-12"] = pop_2023oct + 2 * delta_per_month_b

    all_pop = {**{k: int(round(v)) for k, v in months_2023.items()}, **confirmed}
    idx = pd.period_range("2023-04", "2025-03", freq="M").astype(str)
    return pd.Series({m: all_pop[m] for m in idx}, name="population")


# ────────────────────────────────────────────────────────────────
# データ読み込み
# ────────────────────────────────────────────────────────────────

def load_bs_sheet(path: str, sheet_name: str, month_map: dict) -> pd.DataFrame:
    """
    BS抽出版ファイルの1シートを読み込み、薬効分類×月のDataFrameを返す。
    列: [薬効分類, 年月, 数量]
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    header = rows[HEADER_ROWS - 2]  # row3 (0-indexed: 2)
    month_cols = {col_name: idx for idx, col_name in enumerate(header)
                  if col_name in month_map}

    records = []
    for r in rows[HEADER_ROWS:]:
        cat = r[0]
        if cat is None:
            continue
        for col_name, idx in month_cols.items():
            val = r[idx]
            if isinstance(val, (int, float)):
                records.append({"薬効分類": int(cat), "年月": month_map[col_name], "数量": val})
    return pd.DataFrame(records)


def load_all_data() -> dict:
    """
    rd10/rd11 各3シートを読み込み、設定ごとに統合した DataFame を返す。

    Returns:
        {
          "外来": DataFrame(薬効分類, 年月, 数量),
          "入院": DataFrame(薬効分類, 年月, 数量),
        }
    """
    sheet_map = {
        "外来(院内)": "注射薬 外来 (院内)",
        "外来(院外)": "注射薬 外来 (院外)",
        "入院":      "注射薬 入院",
    }
    file_month_maps = [
        ("rd10", BS_ONLY_FILES["rd10"], FY2023_MONTH_MAP),
        ("rd11", BS_ONLY_FILES["rd11"], FY2024_MONTH_MAP),
    ]

    frames = {"外来(院内)": [], "外来(院外)": [], "入院": []}
    for _label, path, month_map in file_month_maps:
        for setting, sheet_name in sheet_map.items():
            df = load_bs_sheet(path, sheet_name, month_map)
            frames[setting].append(df)

    # 外来院内+院外 を合算
    df_gainai = pd.concat(frames["外来(院内)"]).groupby(["薬効分類", "年月"])["数量"].sum().reset_index()
    df_gaigai = pd.concat(frames["外来(院外)"]).groupby(["薬効分類", "年月"])["数量"].sum().reset_index()
    df_gaigai["数量"] += df_gainai.set_index(["薬効分類", "年月"])["数量"]

    df_gairai = (
        pd.concat([df_gainai, df_gaigai])
        .groupby(["薬効分類", "年月"])["数量"].sum().reset_index()
    )

    df_nyuin = pd.concat(frames["入院"]).groupby(["薬効分類", "年月"])["数量"].sum().reset_index()

    return {"外来": df_gairai, "入院": df_nyuin}


# ────────────────────────────────────────────────────────────────
# ITS モデル
# ────────────────────────────────────────────────────────────────

def build_its_series(df_setting: pd.DataFrame, cat: int, pop: pd.Series) -> pd.DataFrame:
    """
    指定薬効分類のITS用DataFrameを構築する。
    Y = 月次BS処方数量 / 人口 * 100,000
    """
    months = pd.period_range("2023-04", "2025-03", freq="M").astype(str).tolist()
    sub = df_setting[df_setting["薬効分類"] == cat].set_index("年月")["数量"]
    sub = sub.reindex(months, fill_value=0)

    df = pd.DataFrame({"年月": months})
    df["time"] = range(1, 25)
    df["数量"] = sub.values
    df["人口"] = pop.values
    df["Y"] = df["数量"] / df["人口"] * 100_000

    # ITS変数
    t0 = months.index(INTERVENTION_MONTH) + 1  # = 13
    df["D"] = (df["time"] >= t0).astype(int)
    df["t_after"] = np.maximum(df["time"] - (t0 - 1), 0) * df["D"]
    return df


def run_its_model(df: pd.DataFrame):
    """
    セグメント回帰: Y = β0 + β1*time + β2*D + β3*t_after + ε
    """
    X = sm.add_constant(df[["time", "D", "t_after"]])
    model = sm.OLS(df["Y"], X).fit()
    return model


def model_summary(model, cat: int, setting: str) -> dict:
    params = model.params
    ci     = model.conf_int()
    pvals  = model.pvalues
    return {
        "薬効分類": cat,
        "薬効分類名": CAT_NAMES.get(cat, ""),
        "設定": setting,
        "β0_切片":      round(params["const"],  4),
        "β1_介入前傾き": round(params["time"],   4),
        "β2_水準変化":   round(params["D"],      4),
        "β3_傾き変化":   round(params["t_after"],4),
        "β2_CI下限":     round(ci.loc["D", 0],  4),
        "β2_CI上限":     round(ci.loc["D", 1],  4),
        "β3_CI下限":     round(ci.loc["t_after", 0], 4),
        "β3_CI上限":     round(ci.loc["t_after", 1], 4),
        "β2_p値":        round(pvals["D"],       4),
        "β3_p値":        round(pvals["t_after"], 4),
        "R2":            round(model.rsquared,   4),
        "N":             int(model.nobs),
    }


# ────────────────────────────────────────────────────────────────
# 可視化
# ────────────────────────────────────────────────────────────────

def plot_its(df: pd.DataFrame, model, cat: int, setting: str):
    months = df["年月"].tolist()
    x_ticks = range(1, 25)
    t0 = df.loc[df["D"] == 1, "time"].min()  # = 13

    # 適合値
    fitted = model.fittedvalues.values

    # 反事実（介入なし継続）
    params = model.params
    counterfactual = (params["const"]
                      + params["time"] * df["time"])

    fig, ax = plt.subplots(figsize=(11, 5))

    # 観測値
    ax.scatter(df["time"], df["Y"], color="steelblue", zorder=5,
               label="観測値", s=50)

    # 適合線（pre / post）
    pre  = df["time"] < t0
    post = df["time"] >= t0
    ax.plot(df.loc[pre,  "time"], fitted[pre],  color="steelblue", lw=2, label="適合（介入前）")
    ax.plot(df.loc[post, "time"], fitted[post], color="tomato",    lw=2, label="適合（介入後）")

    # 反事実
    ax.plot(df.loc[post, "time"], counterfactual[post], color="steelblue",
            lw=2, ls="--", label="反事実（介入なし推定）")

    # 介入線
    ax.axvline(x=t0 - 0.5, color="gray", ls=":", lw=1.5, label="介入（2024年4月）")

    # X軸ラベル（3ヶ月おき）
    tick_pos   = [1, 4, 7, 10, 13, 16, 19, 22, 24]
    tick_label = [months[i - 1][2:] for i in tick_pos]
    ax.set_xticks(tick_pos)
    ax.set_xticklabels(tick_label, fontsize=8)

    β2 = params["D"]
    β3 = params["t_after"]
    p2 = model.pvalues["D"]
    p3 = model.pvalues["t_after"]
    ax.set_title(
        f"ITS: 薬効分類{cat} {CAT_NAMES.get(cat,'')}  [{setting}]\n"
        f"β2(水準変化)={β2:.2f} (p={p2:.3f})  β3(傾き変化)={β3:.2f} (p={p3:.3f})",
        fontsize=10
    )
    ax.set_xlabel("年月", fontsize=9)
    ax.set_ylabel("BS処方数量（人口100,000人あたり）", fontsize=9)
    ax.legend(fontsize=8, loc="upper left")
    ax.grid(axis="y", alpha=0.3)

    fig.tight_layout()
    fname = OUTPUT_DIR / f"its_{cat}_{setting}.png"
    fig.savefig(fname, dpi=150)
    plt.close(fig)
    return fname


# ────────────────────────────────────────────────────────────────
# メイン
# ────────────────────────────────────────────────────────────────

def main():
    print("データ読み込み中...")
    data   = load_all_data()
    pop    = build_population_series()
    cats   = sorted(CAT_NAMES.keys())
    settings = ["外来", "入院"]

    results = []
    plots   = []

    for setting in settings:
        df_setting = data[setting]
        for cat in cats:
            df_its = build_its_series(df_setting, cat, pop)
            if df_its["Y"].sum() == 0:
                print(f"  SKIP (データなし): 分類{cat} [{setting}]")
                continue

            model = run_its_model(df_its)
            results.append(model_summary(model, cat, setting))
            fname = plot_its(df_its, model, cat, setting)
            plots.append(fname)
            β2p = model.pvalues["D"]
            β3p = model.pvalues["t_after"]
            sig = "**" if min(β2p, β3p) < 0.05 else "  "
            print(f"  {sig} 分類{cat} [{setting}]  β2={model.params['D']:+.2f}(p={β2p:.3f})  β3={model.params['t_after']:+.2f}(p={β3p:.3f})")

    # Excel 出力
    results_df = pd.DataFrame(results)
    out_xlsx = OUTPUT_DIR / "its_results.xlsx"
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        results_df.to_excel(writer, sheet_name="ITS結果", index=False)

        # 書式設定
        ws = writer.sheets["ITS結果"]
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill("solid", fgColor="4472C4")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # p<0.05 の行をハイライト
        sig_fill = PatternFill("solid", fgColor="FFF2CC")
        for row in ws.iter_rows(min_row=2):
            try:
                p2 = float(row[12].value)  # β2_p値列
                p3 = float(row[13].value)  # β3_p値列
                if p2 < 0.05 or p3 < 0.05:
                    for cell in row:
                        cell.fill = sig_fill
            except (TypeError, ValueError):
                pass

        # 列幅調整
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 25)

    print(f"\n結果Excel: {out_xlsx}")
    print(f"グラフ保存先: {OUTPUT_DIR}/its_*.png")
    print(f"合計モデル数: {len(results)}")


if __name__ == "__main__":
    # 日本語フォント設定（Windows Fonts から優先して使用）
    import matplotlib.font_manager as fm
    _JP_CANDIDATES = [
        "/mnt/c/Windows/Fonts/YuGothB.ttc",
        "/mnt/c/Windows/Fonts/meiryo.ttc",
        "/mnt/c/Windows/Fonts/msgothic.ttc",
    ]
    for _fp in _JP_CANDIDATES:
        if Path(_fp).exists():
            fm.fontManager.addfont(_fp)
            plt.rcParams["font.family"] = fm.FontProperties(fname=_fp).get_name()
            break

    main()
