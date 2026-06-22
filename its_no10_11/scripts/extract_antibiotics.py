#!/usr/bin/env python3
"""
Extract antibiotic rows from NDB monthly injection data.
Creates highlighted version and antibiotics-only version.

Usage:
    uv run scripts/extract_antibiotics.py <input.xlsx> <output_dir> [--label rd10|rd11]
"""
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import load_workbook, Workbook
import argparse
import sys
from pathlib import Path

# ── 薬効分類コード設定（変数化） ──────────────────────────────────────
ANTIBIOTIC_CODES: set[int] = {
    611, 612, 613, 614, 615,  # 抗生物質製剤各種
    616,                       # 抗酸菌
    617,                       # 抗真菌（カビ）※スライドから変更して含める
    619,                       # その他の抗生物質製剤
    621,                       # サルファ剤
    622,                       # 抗結核剤
    623,                       # 抗ハンセン病剤
    624,                       # 合成抗菌剤
    629,                       # その他化学療法剤（下記 EXCLUDE_629 を除く）
    # 625 抗ウイルス → 除外
}

# 629 内で除外する医薬品コード（抗ウイルス薬）
EXCLUDE_629_DRUG_CODES: set[int] = {
    622136101,  # サムチレール（Saquinavir、抗HIV薬）
}

# ハイライト色
HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

HEADER_ROWS = 4  # 行0-3はヘッダー（データは行4以降）


def is_antibiotic_row(code: int | None, drug_code: int | None) -> bool | None:
    """
    Returns True if row is antibiotic, False if not, None if code unknown.
    """
    if code is None:
        return None
    if code not in ANTIBIOTIC_CODES:
        return False
    if code == 629 and drug_code in EXCLUDE_629_DRUG_CODES:
        return False
    return True


def create_highlighted(input_path: Path, output_path: Path) -> None:
    """Yellow-highlight all antibiotic rows in the workbook."""
    wb = load_workbook(input_path)
    ws = wb.active

    current_code: int | None = None
    current_is_ab: bool = False

    for row_idx, row in enumerate(ws.iter_rows()):
        if row_idx < HEADER_ROWS:
            continue

        raw_code = row[0].value
        if raw_code is not None and isinstance(raw_code, (int, float)):
            current_code = int(raw_code)
            drug_code = row[2].value
            result = is_antibiotic_row(current_code, drug_code)
            current_is_ab = bool(result) if result is not None else current_is_ab
        else:
            # None in col A → same drug class as previous row; inherit is_ab
            if raw_code is None:
                drug_code = row[2].value
                result = is_antibiotic_row(current_code, drug_code)
                if result is not None:
                    current_is_ab = bool(result)

        if current_is_ab:
            for cell in row:
                cell.fill = HIGHLIGHT_FILL

    wb.save(output_path)
    print(f"  Saved: {output_path}")


def create_ab_only(input_path: Path, output_path: Path) -> None:
    """Extract antibiotic rows only, forward-filling columns A (薬効分類) and B (薬効分類名称)."""
    src_wb = load_workbook(input_path)
    src_ws = src_wb.active

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = src_ws.title

    current_code: int | None = None
    current_name: str | None = None
    current_is_ab: bool = False

    for row_idx, row in enumerate(src_ws.iter_rows(values_only=True)):
        # Always copy header rows as-is
        if row_idx < HEADER_ROWS:
            new_ws.append(list(row))
            continue

        raw_code = row[0]
        if raw_code is not None and isinstance(raw_code, (int, float)):
            current_code = int(raw_code)
            current_name = row[1]
            result = is_antibiotic_row(current_code, row[2])
            current_is_ab = bool(result) if result is not None else current_is_ab
        else:
            if raw_code is None:
                result = is_antibiotic_row(current_code, row[2])
                if result is not None:
                    current_is_ab = bool(result)

        if current_is_ab:
            filled_row = list(row)
            filled_row[0] = current_code   # forward-fill 薬効分類
            filled_row[1] = current_name   # forward-fill 薬効分類名称
            new_ws.append(filled_row)

    new_wb.save(output_path)
    print(f"  Saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Extract antibiotic rows from NDB injection monthly data")
    parser.add_argument("input", help="Input xlsx path")
    parser.add_argument("output_dir", help="Output directory")
    parser.add_argument("--label", default="rd", help="Label prefix (e.g. rd10, rd11)")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    label = args.label

    print(f"\n[{label}] Input: {input_path}")

    hl_path = output_dir / f"{label}_injection_monthly_highlighted.xlsx"
    ab_path = output_dir / f"{label}_injection_monthly_AB_only.xlsx"

    print(f"  Creating highlighted...")
    create_highlighted(input_path, hl_path)

    print(f"  Creating antibiotics-only (with forward-fill)...")
    create_ab_only(input_path, ab_path)

    print(f"  Done.\n")


if __name__ == "__main__":
    main()
